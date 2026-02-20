"""
Comprehensive tests for xls_date_format_extract.py
100% coverage target
"""

import os
import sys
import tempfile
import pytest
from pathlib import Path
from unittest.mock import patch, MagicMock
import openpyxl
from openpyxl import Workbook

# Add src to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from src.integrations.xls_date_format_extract import detect_excel_format, extract_balance_year


@pytest.fixture
def temp_xlsx_full_format():
    """Create XLSX file with full format signature"""
    wb = Workbook()
    ws = wb.active
    # Row 4 should contain "Crediti verso soci" or "Parte richiamata"
    ws['A4'] = "Crediti verso soci per versamenti ancora dovuti"
    ws['B4'] = "Parte richiamata"
    # Add many rows to trigger row count check
    for i in range(5, 150):
        ws[f'A{i}'] = f"Row {i}"
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        yield tmp.name
    if os.path.exists(tmp.name):
        os.remove(tmp.name)


@pytest.fixture
def temp_xlsx_abbreviated_format():
    """Create XLSX file with abbreviated format signature"""
    wb = Workbook()
    ws = wb.active
    # Row 4 should contain "Immobilizzazioni" but NOT "Crediti verso soci"
    ws['A4'] = "Immobilizzazioni"
    ws['B4'] = "Immobilizzazioni Immateriali"
    # Add few rows
    for i in range(5, 50):
        ws[f'A{i}'] = f"Row {i}"
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        yield tmp.name
    if os.path.exists(tmp.name):
        os.remove(tmp.name)


def test_detect_excel_format_full_by_text(temp_xlsx_full_format):
    """Test detection of full format by text signature"""
    result = detect_excel_format(temp_xlsx_full_format)
    assert result == "full"


def test_detect_excel_format_full_by_row_count():
    """Test detection of full format by row count (>120)"""
    wb = Workbook()
    ws = wb.active
    ws['A4'] = "Some other text"  # Not matching text patterns
    # Add many rows
    for i in range(5, 150):
        ws[f'A{i}'] = f"Row {i}"
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    try:
        result = detect_excel_format(tmp_path)
        assert result == "full"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_detect_excel_format_abbreviated_by_text(temp_xlsx_abbreviated_format):
    """Test detection of abbreviated format by text signature"""
    result = detect_excel_format(temp_xlsx_abbreviated_format)
    assert result == "abbreviated"


def test_detect_excel_format_abbreviated_by_row_count():
    """Test detection of abbreviated format by row count (<=120)"""
    wb = Workbook()
    ws = wb.active
    ws['A4'] = "Some other text"  # Not matching text patterns
    # Add few rows
    for i in range(5, 50):
        ws[f'A{i}'] = f"Row {i}"
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    try:
        result = detect_excel_format(tmp_path)
        assert result == "abbreviated"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_detect_excel_format_file_not_found():
    """Test detection with non-existent file"""
    with pytest.raises(FileNotFoundError):
        detect_excel_format("nonexistent_file.xlsx")


def test_detect_excel_format_part_richiamata():
    """Test detection when row 4 contains 'parte richiamata'"""
    wb = Workbook()
    ws = wb.active
    ws['A4'] = "Parte richiamata"
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    try:
        result = detect_excel_format(tmp_path)
        assert result == "full"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_extract_balance_year_date_format_dd_mm_yyyy():
    """Test extraction of date in dd/mm/yyyy format"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Bilancio al 31/12/2023"
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    try:
        result = extract_balance_year(tmp_path)
        assert result == "2023-12-31"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_extract_balance_year_date_format_dd_mm_yyyy_dash():
    """Test extraction of date in dd-mm-yyyy format"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Bilancio al 31-12-2023"
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    try:
        result = extract_balance_year(tmp_path)
        assert result == "2023-12-31"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_extract_balance_year_date_format_dd_mm_yyyy_dot():
    """Test extraction of date in dd.mm.yyyy format"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Bilancio al 31.12.2023"
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    try:
        result = extract_balance_year(tmp_path)
        assert result == "2023-12-31"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_extract_balance_year_year_only():
    """Test extraction of year only (4 digits)"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Bilancio 2023"
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    try:
        result = extract_balance_year(tmp_path)
        assert result == "2023"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_extract_balance_year_italian_month_format():
    """Test extraction with Italian month name"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Bilancio al 31 dicembre 2023"
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    try:
        result = extract_balance_year(tmp_path)
        assert result == "2023-12-31"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_extract_balance_year_italian_month_all_months():
    """Test extraction with all Italian month names"""
    months = [
        "gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
        "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre"
    ]
    
    for i, month in enumerate(months, 1):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = f"Bilancio al 15 {month} 2023"
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        try:
            result = extract_balance_year(tmp_path)
            assert result == f"2023-{i:02d}-15"
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)


def test_extract_balance_year_italian_month_no_day():
    """Test extraction with Italian month but no day specified"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Bilancio dicembre 2023"
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    try:
        result = extract_balance_year(tmp_path)
        assert result == "2023-12-01"  # Defaults to day 01
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_extract_balance_year_no_match():
    """Test extraction when no date pattern matches"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Some text without date"
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    try:
        result = extract_balance_year(tmp_path)
        assert result is None
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_extract_balance_year_empty_file():
    """Test extraction with empty file"""
    wb = Workbook()
    ws = wb.active
    # No content
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    try:
        result = extract_balance_year(tmp_path)
        assert result is None
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_extract_balance_year_invalid_date_format():
    """Test extraction with invalid date format (should return year only)"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Bilancio 2023-13-45"  # Invalid date
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    try:
        result = extract_balance_year(tmp_path)
        # Should return the year part or None
        assert result is not None
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_extract_balance_year_multiple_dates():
    """Test extraction when multiple dates are present (should return first match)"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Bilancio 2022"
    ws['A2'] = "Bilancio 2023"
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    try:
        result = extract_balance_year(tmp_path)
        # Should return first match (2022)
        assert result == "2022"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_extract_balance_year_context_pattern():
    """Test extraction with context patterns like 'esercizio', 'bilancio', 'al'"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Esercizio 2023"
    ws['A2'] = "Bilancio 2024"
    ws['A3'] = "Al 31 dicembre 2025"
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    try:
        result = extract_balance_year(tmp_path)
        # Should find one of the years
        assert result is not None
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_extract_balance_year_file_not_found():
    """Test extraction with non-existent file"""
    with pytest.raises(FileNotFoundError):
        extract_balance_year("nonexistent_file.xlsx")


def test_extract_balance_year_invalid_file():
    """Test extraction with invalid Excel file"""
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp.write(b"not an excel file")
        tmp.flush()
        tmp_path = tmp.name
    try:
        with pytest.raises(Exception):
            extract_balance_year(tmp_path)
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_detect_excel_format_case_insensitive():
    """Test that detection is case insensitive"""
    wb = Workbook()
    ws = wb.active
    ws['A4'] = "CREDITI VERSO SOCI"  # Uppercase
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    try:
        result = detect_excel_format(tmp_path)
        assert result == "full"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_extract_balance_year_multiple_cells():
    """Test extraction when date is spread across multiple cells"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Bilancio"
    ws['B1'] = "al"
    ws['C1'] = "31/12/2023"
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    try:
        result = extract_balance_year(tmp_path)
        assert result == "2023-12-31"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

