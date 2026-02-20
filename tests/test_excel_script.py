"""
Comprehensive tests for excel_script.py (Full format Excel extraction)
100% coverage target
"""

import os
import sys
import tempfile
import pytest
from pathlib import Path
from unittest.mock import patch, mock_open, MagicMock
import openpyxl
from openpyxl import Workbook

# Add src to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from src.integrations.excel_script import extract_bilancio_from_xlsx, main


@pytest.fixture
def temp_xlsx_file():
    """Create a temporary XLSX file with test data for full format"""
    wb = Workbook()
    ws = wb.active
    
    # Set up the structure for full format (ORDINARIO)
    # Row 4: Crediti verso soci - Parte richiamata
    ws['K4'] = 1000.0
    ws['K5'] = 2000.0  # Parte da richiamare
    ws['K6'] = 3000.0  # Totale
    
    # Row 9-16: Immobilizzazioni Immateriali
    ws['K9'] = 500.0   # Costi impianto
    ws['K10'] = 600.0  # Costi sviluppo
    ws['K11'] = 700.0  # Diritti brevetto
    ws['K12'] = 800.0  # Concessioni
    ws['K13'] = 900.0  # Avviamento
    ws['K14'] = 100.0  # Immobilizzazioni in corso
    ws['K15'] = 200.0  # Altre
    ws['K16'] = 3800.0  # Totale
    
    # Row 18-23: Immobilizzazioni Materiali
    ws['K18'] = 10000.0  # Terreni e fabbricati
    ws['K19'] = 20000.0  # Impianti e macchinari
    ws['K20'] = 5000.0   # Attrezzature
    ws['K21'] = 3000.0   # Altri beni
    ws['K22'] = 2000.0   # Immobilizzazioni in corso
    ws['K23'] = 40000.0  # Totale
    
    # Row 26-31: Partecipazioni
    ws['K26'] = 5000.0  # Imprese controllate
    ws['K27'] = 3000.0  # Imprese collegate
    ws['K28'] = 2000.0  # Imprese controllanti
    ws['K29'] = 1000.0  # Imprese sottoposte
    ws['K30'] = 500.0   # Altre imprese
    ws['K31'] = 11500.0  # Totale
    
    # Row 34-53: Crediti nelle immobilizzazioni finanziarie
    ws['K34'] = 1000.0  # Verso imprese controllate - esigibili entro
    ws['K35'] = 2000.0  # Verso imprese controllate - esigibili oltre
    ws['K36'] = 3000.0  # Totale verso controllate
    ws['K38'] = 500.0   # Verso imprese collegate - esigibili entro
    ws['K39'] = 1000.0  # Verso imprese collegate - esigibili oltre
    ws['K40'] = 1500.0  # Totale verso collegate
    ws['K42'] = 300.0   # Verso imprese controllanti - esigibili entro
    ws['K43'] = 700.0   # Verso imprese controllanti - esigibili oltre
    ws['K44'] = 1000.0  # Totale verso controllanti
    ws['K46'] = 200.0   # Imprese sottoposte - esigibili entro
    ws['K47'] = 300.0   # Imprese sottoposte - esigibili oltre
    ws['K48'] = 500.0   # Totale verso sottoposte
    ws['K50'] = 100.0   # Altre imprese - esigibili entro
    ws['K51'] = 200.0   # Altre imprese - esigibili oltre
    ws['K52'] = 300.0   # Totale verso altre
    ws['K53'] = 6300.0  # Totale crediti
    ws['K54'] = 500.0   # Altri titoli
    ws['K55'] = 100.0   # Strumenti finanziari derivati attivi
    ws['K56'] = 23400.0  # Totale immobilizzazioni
    
    # Row 59-64: Rimanenze
    ws['K59'] = 2000.0  # Materie prime
    ws['K60'] = 1500.0  # Prodotti in corso
    ws['K61'] = 1000.0  # Lavori in corso
    ws['K62'] = 3000.0  # Prodotti finiti
    ws['K63'] = 500.0   # Acconti
    ws['K64'] = 8000.0  # Totale rimanenze
    
    # Row 67-95: Crediti
    ws['K67'] = 5000.0  # Verso clienti - esigibili entro
    ws['K68'] = 2000.0  # Verso clienti - esigibili oltre
    ws['K69'] = 7000.0  # Totale verso clienti
    ws['K71'] = 1000.0  # Verso imprese controllate - esigibili entro
    ws['K72'] = 500.0   # Verso imprese controllate - esigibili oltre
    ws['K73'] = 1500.0  # Totale verso controllate
    ws['K75'] = 800.0   # Verso imprese collegate - esigibili entro
    ws['K76'] = 200.0   # Verso imprese collegate - esigibili oltre
    ws['K77'] = 1000.0  # Totale verso collegate
    ws['K79'] = 300.0   # Verso Controllanti - esigibili entro
    ws['K80'] = 100.0   # Verso Controllanti - esigibili oltre
    ws['K81'] = 400.0   # Totale verso controllanti
    ws['K83'] = 200.0   # Verso imprese sottoposte - esigibili entro
    ws['K84'] = 50.0    # Verso imprese sottoposte - esigibili oltre
    ws['K85'] = 250.0   # Totale verso sottoposte
    ws['K87'] = 500.0   # Crediti tributari - esigibili entro
    ws['K88'] = 100.0   # Crediti tributari - esigibili oltre
    ws['K89'] = 600.0   # Totale crediti tributari
    ws['K90'] = 200.0   # Imposte anticipate
    ws['K92'] = 300.0   # Verso altri - esigibili entro
    ws['K93'] = 100.0   # Verso altri - esigibili oltre
    ws['K94'] = 400.0   # Totale verso altri
    ws['K95'] = 11150.0  # Totale crediti
    
    # Row 97-104: Attività finanziarie
    ws['K97'] = 1000.0  # Partecipazioni in imprese controllate
    ws['K98'] = 500.0   # Partecipazioni in imprese collegate
    ws['K99'] = 300.0   # Partecipazioni in imprese controllanti
    ws['K100'] = 200.0  # Partecipazioni in imprese sottoposte
    ws['K101'] = 100.0  # Altre partecipazioni
    ws['K102'] = 50.0   # Strumenti finanziari derivati attivi
    ws['K103'] = 150.0  # Altri titoli
    ws['K104'] = 2300.0  # Totale attività finanziarie
    
    # Row 106-109: Disponibilità liquide
    ws['K106'] = 5000.0  # Depositi bancari
    ws['K107'] = 200.0   # Assegni
    ws['K108'] = 100.0   # Denaro e valori in cassa
    ws['K109'] = 5300.0  # Totale disponibilità liquide
    ws['K109'] = 5300.0  # Totale attivo circolante (same row)
    
    # Row 111-112: Ratei e risconti attivi
    ws['K111'] = 500.0   # Ratei attivi
    ws['K112'] = 200.0   # Ratei inattivi
    ws['K113'] = 50000.0  # Totale attivo
    
    # PASSIVO - Row 116-141: Patrimonio Netto
    ws['K116'] = 10000.0  # Capitale
    ws['K117'] = 2000.0   # Riserve da sovrapprezzo azioni
    ws['K118'] = 1500.0   # Riserve da rivalutazioni
    ws['K119'] = 1000.0   # Riserve legali
    ws['K120'] = 500.0    # Riserve statutarie
    ws['K122'] = 300.0    # Riserva straordinaria
    ws['K123'] = 200.0    # Riserva da deroghe
    ws['K124'] = 100.0    # Riserva azioni controllante
    ws['K125'] = 150.0    # Riserva da rivalutazione partecipazioni
    ws['K126'] = 50.0     # Versamenti in conto aumento capitale
    ws['K127'] = 25.0     # Versamenti in conto futuro aumento
    ws['K128'] = 10.0     # Versamenti in conto capitale
    ws['K129'] = 5.0      # Versamenti a copertura perdite
    ws['K130'] = 20.0     # Riserva da riduzione capitale
    ws['K131'] = 30.0     # Riserva avanzo di fusione
    ws['K132'] = 40.0     # Riserva per utili su cambi
    ws['K133'] = 15.0     # Riserva da conguaglio utili
    ws['K134'] = 25.0     # Varie altre riserve
    ws['K135'] = 995.0    # Totale altre riserve
    ws['K136'] = 100.0    # Riserve per operazioni copertura
    ws['K137'] = 500.0    # Utili/Perdite portati a nuovo
    ws['K138'] = 2000.0   # Utile/Perdita dell'esercizio
    ws['K139'] = 0.0      # Perdita ripristinata
    ws['K140'] = 0.0      # Riserva negativa azioni proprie
    ws['K141'] = 18620.0  # Totale patrimonio netto
    
    # Row 143-147: Fondi per rischi e oneri
    ws['K143'] = 500.0   # Per trattamento quiescenza
    ws['K144'] = 300.0   # Per imposte anche differite
    ws['K145'] = 100.0   # Strumenti finanziari derivati passivi
    ws['K146'] = 200.0   # Altri
    ws['K147'] = 1100.0  # Totale fondi
    
    # Row 148: TFR
    ws['K148'] = 2000.0  # Trattamento fine rapporto
    
    # Row 151-210: Debiti
    ws['K151'] = 1000.0  # Obbligazioni - esigibili entro
    ws['K152'] = 2000.0  # Obbligazioni - esigibili oltre
    ws['K153'] = 3000.0  # Totale obbligazioni
    ws['K155'] = 500.0   # Obbligazioni convertibili - esigibili entro
    ws['K156'] = 1000.0  # Obbligazioni convertibili - esigibili oltre
    ws['K157'] = 1500.0  # Totale obbligazioni convertibili
    ws['K159'] = 200.0   # Debiti verso soci - esigibili entro
    ws['K160'] = 300.0   # Debiti verso soci - esigibili oltre
    ws['K161'] = 500.0   # Totale debiti verso soci
    ws['K163'] = 5000.0  # Debiti verso banche - esigibili entro
    ws['K164'] = 3000.0  # Debiti verso banche - esigibili oltre
    ws['K165'] = 8000.0  # Totale debiti verso banche
    ws['K167'] = 1000.0  # Debiti verso altri finanziatori - esigibili entro
    ws['K168'] = 500.0   # Debiti verso altri finanziatori - esigibili oltre
    ws['K169'] = 1500.0  # Totale debiti verso altri finanziatori
    ws['K171'] = 800.0   # Acconti - esigibili entro
    ws['K172'] = 200.0   # Acconti - esigibili oltre
    ws['K173'] = 1000.0  # Totale acconti
    ws['K175'] = 3000.0  # Debiti verso fornitori - esigibili entro
    ws['K176'] = 1000.0  # Debiti verso fornitori - esigibili oltre
    ws['K177'] = 4000.0  # Totale debiti verso fornitori
    ws['K179'] = 200.0   # Debiti rappresentati da titoli - esigibili entro
    ws['K180'] = 100.0   # Debiti rappresentati da titoli - esigibili oltre
    ws['K181'] = 300.0   # Totale debiti rappresentati da titoli
    ws['K183'] = 500.0   # Debiti verso imprese controllate - esigibili entro
    ws['K184'] = 200.0   # Debiti verso imprese controllate - esigibili oltre
    ws['K185'] = 700.0   # Totale debiti verso controllate
    ws['K187'] = 300.0   # Debiti verso imprese collegate - esigibili entro
    ws['K188'] = 100.0   # Debiti verso imprese collegate - esigibili oltre
    ws['K189'] = 400.0   # Totale debiti verso collegate
    ws['K191'] = 200.0   # Debiti verso controllanti - esigibili entro
    ws['K192'] = 50.0    # Debiti verso controllanti - esigibili oltre
    ws['K193'] = 250.0   # Totale debiti verso controllanti
    ws['K195'] = 100.0   # Debiti verso imprese sottoposte - esigibili entro
    ws['K196'] = 25.0    # Debiti verso imprese sottoposte - esigibili oltre
    ws['K197'] = 125.0   # Totale debiti verso sottoposte
    ws['K199'] = 400.0   # Debiti tributari - esigibili entro
    ws['K200'] = 100.0   # Debiti tributari - esigibili oltre
    ws['K201'] = 500.0   # Totale debiti tributari
    ws['K203'] = 300.0   # Debiti verso istituti previdenza - esigibili entro
    ws['K204'] = 50.0    # Debiti verso istituti previdenza - esigibili oltre
    ws['K205'] = 350.0   # Totale debiti verso istituti
    ws['K207'] = 200.0   # Altri debiti - esigibili entro
    ws['K208'] = 50.0    # Altri debiti - esigibili oltre
    ws['K209'] = 250.0   # Totale altri debiti
    ws['K210'] = 20125.0  # Totale debiti
    
    # Row 212-213: Ratei e risconti passivi
    ws['K212'] = 300.0   # Ratei passivi
    ws['K213'] = 100.0   # Risconti passivi
    ws['K214'] = 50000.0  # Totale passivo
    
    # CONTO ECONOMICO - Row 217-309
    ws['K217'] = 100000.0  # Ricavi vendite
    ws['K218'] = 1000.0    # Variazione lavorazioni in corso
    ws['K219'] = 500.0     # Variazione lavori in corso
    ws['K220'] = 200.0     # Incrementi immobilizzazioni
    ws['K222'] = 500.0     # Contributi in conto esercizio
    ws['K223'] = 300.0     # Altri
    ws['K224'] = 800.0     # Totale altri ricavi
    ws['K225'] = 102500.0  # Totale valore produzione
    
    # Costi di produzione
    ws['K227'] = 30000.0  # Materie prime
    ws['K228'] = 15000.0  # Servizi
    ws['K229'] = 5000.0   # Godimento beni terzi
    ws['K231'] = 20000.0  # Salari e stipendi
    ws['K232'] = 5000.0   # Oneri sociali
    ws['K233'] = 1000.0   # TFR
    ws['K234'] = 500.0    # Trattamento quiescenza
    ws['K235'] = 500.0    # Altri costi
    ws['K236'] = 31000.0  # Totale costi personale
    ws['K238'] = 2000.0   # Ammortamento immateriali
    ws['K239'] = 5000.0   # Ammortamento materiali
    ws['K240'] = 500.0    # Altre svalutazioni
    ws['K241'] = 300.0    # Svalutazioni crediti
    ws['K242'] = 7800.0   # Totale ammortamenti
    ws['K243'] = 1000.0   # Variazione rimanenze materie prime
    ws['K244'] = 500.0    # Accantonamento per rischi
    ws['K245'] = 300.0    # Altri accantonamenti
    ws['K246'] = 2000.0   # Oneri diversi
    ws['K247'] = 85600.0  # Totale costi produzione
    ws['K248'] = 16900.0  # Differenza valore produzione - costi produzione
    
    # Proventi e oneri finanziari
    ws['K251'] = 1000.0  # Proventi da partecipazioni - controllate
    ws['K252'] = 500.0   # Proventi da partecipazioni - collegate
    ws['K253'] = 300.0   # Proventi da partecipazioni - controllanti
    ws['K254'] = 200.0   # Proventi da partecipazioni - sottoposte
    ws['K255'] = 100.0   # Proventi da partecipazioni - altre
    ws['K256'] = 2100.0  # Totale proventi partecipazioni
    ws['K259'] = 500.0   # Proventi da crediti - controllate
    ws['K260'] = 300.0   # Proventi da crediti - collegate
    ws['K261'] = 200.0   # Proventi da crediti - controllanti
    ws['K262'] = 100.0   # Proventi da crediti - sottoposte
    ws['K263'] = 50.0    # Proventi da crediti - altri
    ws['K264'] = 1150.0  # Totale proventi da crediti
    ws['K265'] = 200.0   # Proventi da titoli immobilizzazioni
    ws['K266'] = 100.0   # Proventi da crediti attivo circolante
    ws['K268'] = 300.0   # Proventi diversi - controllate
    ws['K269'] = 200.0   # Proventi diversi - collegate
    ws['K270'] = 100.0   # Proventi diversi - controllanti
    ws['K271'] = 50.0    # Proventi diversi - sottoposte
    ws['K272'] = 25.0    # Proventi diversi - altri
    ws['K273'] = 675.0   # Totale proventi diversi
    ws['K274'] = 2225.0  # Totale altri proventi finanziari
    ws['K276'] = 1000.0  # Interessi e oneri - controllate
    ws['K277'] = 500.0   # Interessi e oneri - collegate
    ws['K278'] = 300.0   # Interessi e oneri - controllanti
    ws['K279'] = 200.0   # Interessi e oneri - sottoposte
    ws['K280'] = 100.0   # Interessi e oneri - altri
    ws['K281'] = 2100.0  # Totale interessi e oneri
    ws['K282'] = 50.0    # Utili e perdite su cambi
    ws['K283'] = 2275.0  # Totale proventi e oneri finanziari
    
    # Rettifiche di valore
    ws['K286'] = 100.0   # Rivalutazioni - partecipazioni
    ws['K287'] = 50.0    # Rivalutazioni - immobilizzazioni finanziarie
    ws['K288'] = 25.0    # Rivalutazioni - titoli attivo circolante
    ws['K289'] = 10.0    # Rivalutazioni - strumenti derivati
    ws['K292'] = 200.0   # Svalutazioni - partecipazioni
    ws['K293'] = 100.0   # Svalutazioni - immobilizzazioni finanziarie
    ws['K294'] = 50.0    # Svalutazioni - titoli attivo circolante
    ws['K295'] = 25.0    # Svalutazioni - strumenti derivati
    ws['K297'] = -275.0  # Totale rettifiche
    
    # Proventi e oneri straordinari
    ws['K299'] = 500.0   # Proventi straordinari
    ws['K300'] = 200.0   # Oneri straordinari
    ws['K301'] = 300.0   # Totale partite straordinarie
    
    # Imposte
    ws['K304'] = 4000.0  # Imposte correnti
    ws['K305'] = 100.0   # Imposte esercizi precedenti
    ws['K306'] = 200.0   # Imposte differite
    ws['K307'] = 50.0    # Proventi da consolidato fiscale
    ws['K308'] = 4350.0  # Totale imposte
    ws['K309'] = 14500.0  # Utile/Perdita dell'esercizio
    
    # Create temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        yield tmp.name
    # Cleanup
    if os.path.exists(tmp.name):
        os.remove(tmp.name)


def test_extract_bilancio_from_xlsx_success(temp_xlsx_file):
    """Test successful extraction of full format Excel file"""
    result = extract_bilancio_from_xlsx(temp_xlsx_file)
    
    assert result is not None
    assert "informazioni_generali" in result
    assert "Stato_patrimoniale" in result
    assert "Conto_economico" in result
    
    # Check some key values
    assert result["Stato_patrimoniale"]["Attivo"]["Crediti_verso_soci_per_versamenti_ancora_dovuti"]["Parte_richiamata"] == 1000.0
    assert result["Stato_patrimoniale"]["Attivo"]["Totale_attivo"] == 50000.0
    assert result["Stato_patrimoniale"]["Passivo"]["Totale_passivo"] == 50000.0
    assert result["Conto_economico"]["Valore_della_produzione"]["Totale_valore_della_produzione"] == 102500.0


def test_extract_bilancio_from_xlsx_empty_cells(temp_xlsx_file):
    """Test extraction with empty cells (should return 0.0)"""
    # Modify file to have empty cells
    wb = openpyxl.load_workbook(temp_xlsx_file)
    ws = wb.active
    ws['K4'] = None
    ws['K5'] = ''
    wb.save(temp_xlsx_file)
    
    result = extract_bilancio_from_xlsx(temp_xlsx_file)
    assert result["Stato_patrimoniale"]["Attivo"]["Crediti_verso_soci_per_versamenti_ancora_dovuti"]["Parte_richiamata"] == 0.0
    assert result["Stato_patrimoniale"]["Attivo"]["Crediti_verso_soci_per_versamenti_ancora_dovuti"]["Parte_da_richiamare"] == 0.0


def test_extract_bilancio_from_xlsx_italian_number_format():
    """Test extraction with Italian number format (comma as decimal separator)"""
    wb = Workbook()
    ws = wb.active
    ws['K4'] = "1.234,56"  # Italian format: 1234.56
    ws['K5'] = "2.345,67"
    
    # On Windows, NamedTemporaryFile keeps the file handle open inside the
    # context manager, so we capture the path and close the file before
    # passing it to openpyxl / removing it.
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    try:
        result = extract_bilancio_from_xlsx(tmp_path)
        assert result["Stato_patrimoniale"]["Attivo"]["Crediti_verso_soci_per_versamenti_ancora_dovuti"]["Parte_richiamata"] == 1234.56
        assert result["Stato_patrimoniale"]["Attivo"]["Crediti_verso_soci_per_versamenti_ancora_dovuti"]["Parte_da_richiamare"] == 2345.67
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_extract_bilancio_from_xlsx_invalid_number_format():
    """Test extraction with invalid number format (should return 0.0)"""
    wb = Workbook()
    ws = wb.active
    ws['K4'] = "not a number"
    ws['K5'] = "abc123"
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    try:
        result = extract_bilancio_from_xlsx(tmp_path)
        assert result["Stato_patrimoniale"]["Attivo"]["Crediti_verso_soci_per_versamenti_ancora_dovuti"]["Parte_richiamata"] == 0.0
        assert result["Stato_patrimoniale"]["Attivo"]["Crediti_verso_soci_per_versamenti_ancora_dovuti"]["Parte_da_richiamare"] == 0.0
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_extract_bilancio_from_xlsx_file_not_found():
    """Test extraction with non-existent file"""
    with pytest.raises(FileNotFoundError):
        extract_bilancio_from_xlsx("nonexistent_file.xlsx")


def test_extract_bilancio_from_xlsx_invalid_file():
    """Test extraction with invalid Excel file"""
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp.write(b"not an excel file")
        tmp.flush()

        tmp_path = tmp.name

    try:
        with pytest.raises(Exception):  # openpyxl will raise an exception
            extract_bilancio_from_xlsx(tmp_path)
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_main_function_success(temp_xlsx_file, capsys):
    """Test main function with valid file"""
    with patch('sys.argv', ['script.py', temp_xlsx_file]):
        try:
            main()
            captured = capsys.readouterr()
            assert "Extraction completed successfully" in captured.out
            assert "Ordinario_Annuale.json" in captured.out
        finally:
            # Clean up JSON artefact created by main() so it doesn't pollute the repo
            if os.path.exists("Ordinario_Annuale.json"):
                os.remove("Ordinario_Annuale.json")


def test_main_function_no_args(capsys):
    """Test main function without arguments"""
    with patch('sys.argv', ['script.py']):
        main()
        captured = capsys.readouterr()
        assert "Usage:" in captured.out


def test_main_function_file_not_found(capsys):
    """Test main function with non-existent file"""
    with patch('sys.argv', ['script.py', 'nonexistent.xlsx']):
        main()
        captured = capsys.readouterr()
        assert "not found" in captured.out.lower()


def test_main_function_extraction_error(capsys):
    """Test main function when extraction fails"""
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp.write(b"invalid excel")
        tmp.flush()
        tmp_path = tmp.name

    try:
        with patch('sys.argv', ['script.py', tmp_path]):
            main()
            captured = capsys.readouterr()
            # Check combined output (stdout + stderr) for maximum robustness
            # This works regardless of where error messages are printed
            combined_output = captured.out + captured.err
            assert "Traceback" in combined_output or len(captured.err) > 0
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_extract_bilancio_structure_completeness(temp_xlsx_file):
    """Test that all expected structure fields are present"""
    result = extract_bilancio_from_xlsx(temp_xlsx_file)
    
    # Check Stato Patrimoniale structure
    assert "Attivo" in result["Stato_patrimoniale"]
    assert "Passivo" in result["Stato_patrimoniale"]
    
    # Check Immobilizzazioni structure
    assert "Immobilizzazioni_Immateriali" in result["Stato_patrimoniale"]["Attivo"]["Immobilizzazioni"]
    assert "Immobilizzazioni_Materiali" in result["Stato_patrimoniale"]["Attivo"]["Immobilizzazioni"]
    assert "Immobilizzazioni_Finanziarie" in result["Stato_patrimoniale"]["Attivo"]["Immobilizzazioni"]
    
    # Check Conto Economico structure
    assert "Valore_della_produzione" in result["Conto_economico"]
    assert "Costi_di_produzione" in result["Conto_economico"]
    assert "Proventi_e_oneri_finanziari" in result["Conto_economico"]


def test_extract_bilancio_numeric_values(temp_xlsx_file):
    """Test that numeric values are correctly extracted"""
    result = extract_bilancio_from_xlsx(temp_xlsx_file)
    
    # Test various numeric extractions
    assert isinstance(result["Stato_patrimoniale"]["Attivo"]["Totale_attivo"], float)
    assert isinstance(result["Stato_patrimoniale"]["Passivo"]["Totale_passivo"], float)
    assert isinstance(result["Conto_economico"]["Valore_della_produzione"]["Totale_valore_della_produzione"], float)

