"""
Comprehensive tests for excel_script_2.py (Abbreviated format Excel extraction)
100% coverage target
"""

import os
import sys
import tempfile
import pytest
from pathlib import Path
from unittest.mock import patch, mock_open, MagicMock
import openpyxl
from openpyxl import Workbook

# Add src to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from src.integrations.excel_script_2 import extract_bilancio_abbreviato_from_xlsx, main


@pytest.fixture
def temp_xlsx_abbreviato_file():
    """Create a temporary XLSX file with test data for abbreviated format"""
    wb = Workbook()
    ws = wb.active
    
    # Set up the structure for abbreviated format (ABBREVIATO)
    # Row 4: Immobilizzazioni Immateriali
    ws['K4'] = 1000.0
    # Row 5: Immobilizzazioni Materiali
    ws['K5'] = 5000.0
    # Row 7-8: Immobilizzazioni Finanziarie
    ws['K7'] = 2000.0   # dell'esercizio corrente
    ws['K8'] = 3000.0   # oltre l'esercizio corrente
    
    # Row 10-13: Attivo Circolante
    ws['K10'] = 4000.0  # Rimanenze
    ws['K11'] = 6000.0  # Crediti
    ws['K12'] = 1000.0  # Attività finanziarie
    ws['K13'] = 2000.0  # Disponibilità liquide
    ws['K14'] = 20000.0  # Totale attivo
    
    # PASSIVO - Row 17-27: Patrimonio Netto
    ws['K17'] = 5000.0   # Capitale
    ws['K18'] = 1000.0   # Riserve da sovrapprezzo azioni
    ws['K19'] = 500.0    # Riserve da rivalutazioni
    ws['K20'] = 300.0    # Riserve legali
    ws['K21'] = 200.0    # Riserve statutarie
    ws['K22'] = 100.0    # Altre riserve
    ws['K23'] = 50.0     # Riserve per operazioni copertura
    ws['K24'] = 200.0    # Utili/Perdite portati a nuovo
    ws['K25'] = 1500.0   # Utile/Perdita dell'esercizio
    ws['K26'] = 0.0      # Riserva negativa azioni proprie
    ws['K27'] = 9650.0   # Totale patrimonio netto
    
    # Row 28: Fondi per rischi e oneri
    ws['K28'] = 500.0
    # Row 29: TFR
    ws['K29'] = 1000.0
    # Row 31-32: Debiti
    ws['K31'] = 5000.0   # dell'esercizio corrente
    ws['K32'] = 3000.0   # oltre l'esercizio corrente
    # Row 33: Ratei e risconti
    ws['K33'] = 350.0
    # Row 34: Totale passivo
    ws['K34'] = 20000.0
    
    # CONTO ECONOMICO - Row 37-129
    # Valore della produzione
    ws['K37'] = 50000.0  # Ricavi vendite
    ws['K38'] = 500.0    # Variazione rimanenze prodotti
    ws['K39'] = 300.0    # Variazione lavori in corso
    ws['K40'] = 200.0    # Incrementi immobilizzazioni
    ws['K42'] = 100.0    # Contributi in conto esercizio
    ws['K43'] = 200.0    # Altri
    ws['K44'] = 300.0    # Totale altri ricavi
    ws['K45'] = 51300.0  # Totale valore produzione
    
    # Costi di produzione
    ws['K47'] = 15000.0  # Materie prime
    ws['K48'] = 8000.0   # Servizi
    ws['K49'] = 2000.0   # Godimento beni terzi
    ws['K51'] = 10000.0  # Salari e stipendi
    ws['K52'] = 2500.0   # Oneri sociali
    ws['K53'] = 500.0    # TFR
    ws['K54'] = 200.0    # Trattamento quiescenza
    ws['K55'] = 300.0    # Altri costi
    ws['K56'] = 13500.0  # Totale costi personale
    ws['K58'] = 1000.0   # Ammortamento immateriali
    ws['K59'] = 3000.0   # Ammortamento materiali
    ws['K60'] = 200.0    # Altre svalutazioni
    ws['K61'] = 100.0    # Svalutazioni crediti
    ws['K62'] = 4300.0   # Totale ammortamenti
    ws['K63'] = 500.0    # Variazione rimanenze materie prime
    ws['K64'] = 200.0    # Accantonamento per rischi
    ws['K65'] = 100.0    # Altri accantonamenti
    ws['K66'] = 1000.0   # Oneri diversi
    ws['K67'] = 40400.0  # Totale costi produzione
    ws['K68'] = 10900.0  # Differenza A-B
    
    # Proventi e oneri finanziari
    ws['K71'] = 500.0    # Proventi partecipazioni - controllate
    ws['K72'] = 300.0    # Proventi partecipazioni - collegate
    ws['K73'] = 200.0    # Proventi partecipazioni - controllanti
    ws['K74'] = 100.0    # Proventi partecipazioni - sottoposte
    ws['K75'] = 50.0     # Proventi partecipazioni - altri
    ws['K76'] = 1150.0  # Totale proventi partecipazioni
    ws['K79'] = 200.0    # Altri proventi - crediti - controllate
    ws['K80'] = 150.0    # Altri proventi - crediti - collegate
    ws['K81'] = 100.0    # Altri proventi - crediti - controllanti
    ws['K82'] = 50.0     # Altri proventi - crediti - sottoposte
    ws['K83'] = 25.0     # Altri proventi - crediti - altri
    ws['K84'] = 525.0    # Totale proventi da crediti
    ws['K85'] = 100.0    # Proventi da titoli immobilizzazioni
    ws['K86'] = 50.0     # Proventi da titoli attivo circolante
    ws['K88'] = 150.0    # Proventi diversi - controllate
    ws['K89'] = 100.0    # Proventi diversi - collegate
    ws['K90'] = 50.0     # Proventi diversi - controllanti
    ws['K91'] = 25.0     # Proventi diversi - sottoposte
    ws['K92'] = 10.0     # Proventi diversi - altri
    ws['K93'] = 335.0    # Totale proventi diversi
    ws['K94'] = 1010.0   # Totale altri proventi finanziari
    ws['K96'] = 800.0    # Interessi e oneri - controllate
    ws['K97'] = 400.0    # Interessi e oneri - collegate
    ws['K98'] = 200.0    # Interessi e oneri - controllanti
    ws['K99'] = 100.0    # Interessi e oneri - sottoposte
    ws['K100'] = 50.0    # Interessi e oneri - altri
    ws['K101'] = 1550.0  # Totale interessi e oneri
    ws['K102'] = 25.0    # Utili e perdite su cambi
    ws['K103'] = 1635.0  # Totale proventi e oneri finanziari
    
    # Rettifiche di valore
    ws['K106'] = 50.0    # Rivalutazioni - partecipazioni
    ws['K107'] = 25.0    # Rivalutazioni - immobilizzazioni finanziarie
    ws['K108'] = 10.0    # Rivalutazioni - titoli attivo circolante
    ws['K109'] = 5.0     # Rivalutazioni - strumenti derivati
    ws['K110'] = 90.0    # Totale rivalutazioni
    ws['K112'] = 100.0   # Svalutazioni - partecipazioni
    ws['K113'] = 50.0    # Svalutazioni - immobilizzazioni finanziarie
    ws['K114'] = 25.0    # Svalutazioni - titoli attivo circolante
    ws['K115'] = 10.0    # Svalutazioni - strumenti derivati
    ws['K116'] = 185.0   # Totale svalutazioni
    ws['K117'] = -95.0   # Totale rettifiche
    
    # Proventi e oneri straordinari
    ws['K119'] = 200.0   # Proventi straordinari
    ws['K120'] = 100.0   # Oneri straordinari
    ws['K121'] = 100.0   # Totale partite straordinarie
    
    # Risultato prima delle imposte
    ws['K122'] = 11740.0
    
    # Imposte
    ws['K124'] = 3000.0  # Imposte correnti
    ws['K125'] = 100.0   # Imposte esercizi precedenti
    ws['K126'] = 200.0   # Imposte differite
    ws['K127'] = 50.0    # Proventi da consolidato fiscale
    ws['K128'] = 3350.0  # Totale imposte
    ws['K129'] = 8390.0  # Utile/Perdita dell'esercizio
    
    # Create temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        yield tmp.name
    # Cleanup
    if os.path.exists(tmp.name):
        os.remove(tmp.name)


def test_extract_bilancio_abbreviato_from_xlsx_success(temp_xlsx_abbreviato_file):
    """Test successful extraction of abbreviated format Excel file"""
    result = extract_bilancio_abbreviato_from_xlsx(temp_xlsx_abbreviato_file)
    
    assert result is not None
    assert "informazioni_generali" in result
    assert "Stato_patrimoniale" in result
    assert "Conto_economico" in result
    
    # Check some key values
    assert result["Stato_patrimoniale"]["Attivo"]["Immobilizzazioni"]["Immobilizzazioni_Immateriali"] == 1000.0
    assert result["Stato_patrimoniale"]["Attivo"]["Totale_attivo"] == 20000.0
    assert result["Stato_patrimoniale"]["Passivo"]["Totale_passivo"] == 20000.0
    assert result["Conto_economico"]["Valore_della_produzione"]["Totale_valore_della_produzione"] == 51300.0


def test_extract_bilancio_abbreviato_empty_cells(temp_xlsx_abbreviato_file):
    """Test extraction with empty cells (should return 0.0)"""
    wb = openpyxl.load_workbook(temp_xlsx_abbreviato_file)
    ws = wb.active
    ws['K4'] = None
    ws['K5'] = ''
    wb.save(temp_xlsx_abbreviato_file)
    
    result = extract_bilancio_abbreviato_from_xlsx(temp_xlsx_abbreviato_file)
    assert result["Stato_patrimoniale"]["Attivo"]["Immobilizzazioni"]["Immobilizzazioni_Immateriali"] == 0.0
    assert result["Stato_patrimoniale"]["Attivo"]["Immobilizzazioni"]["Immobilizzazioni_Materiali"] == 0.0


def test_extract_bilancio_abbreviato_italian_number_format():
    """Test extraction with Italian number format"""
    wb = Workbook()
    ws = wb.active
    ws['K4'] = "1.234,56"
    ws['K5'] = "2.345,67"
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    try:
        result = extract_bilancio_abbreviato_from_xlsx(tmp_path)
        assert result["Stato_patrimoniale"]["Attivo"]["Immobilizzazioni"]["Immobilizzazioni_Immateriali"] == 1234.56
        assert result["Stato_patrimoniale"]["Attivo"]["Immobilizzazioni"]["Immobilizzazioni_Materiali"] == 2345.67
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_extract_bilancio_abbreviato_invalid_number_format():
    """Test extraction with invalid number format"""
    wb = Workbook()
    ws = wb.active
    ws['K4'] = "not a number"
    ws['K5'] = "abc123"
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    try:
        result = extract_bilancio_abbreviato_from_xlsx(tmp_path)
        assert result["Stato_patrimoniale"]["Attivo"]["Immobilizzazioni"]["Immobilizzazioni_Immateriali"] == 0.0
        assert result["Stato_patrimoniale"]["Attivo"]["Immobilizzazioni"]["Immobilizzazioni_Materiali"] == 0.0
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_extract_bilancio_abbreviato_file_not_found():
    """Test extraction with non-existent file"""
    with pytest.raises(FileNotFoundError):
        extract_bilancio_abbreviato_from_xlsx("nonexistent_file.xlsx")


def test_extract_bilancio_abbreviato_invalid_file():
    """Test extraction with invalid Excel file"""
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp.write(b"not an excel file")
        tmp.flush()
        tmp_path = tmp.name

    try:
        with pytest.raises(Exception):
            extract_bilancio_abbreviato_from_xlsx(tmp_path)
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_main_function_success(temp_xlsx_abbreviato_file, capsys):
    """Test main function with valid file"""
    with patch('sys.argv', ['script.py', temp_xlsx_abbreviato_file]):
        try:
            main()
            captured = capsys.readouterr()
            assert "Extraction completed successfully" in captured.out
            assert "Abbreviato_Extracted_3.json" in captured.out
        finally:
            # Clean up JSON artefact created by main() so it doesn't pollute the repo
            if os.path.exists("Abbreviato_Extracted_3.json"):
                os.remove("Abbreviato_Extracted_3.json")


def test_main_function_no_args(capsys):
    """Test main function without arguments"""
    with patch('sys.argv', ['script.py']):
        main()
        captured = capsys.readouterr()
        assert "Usage:" in captured.out


def test_main_function_file_not_found(capsys):
    """Test main function with non-existent file"""
    with patch('sys.argv', ['script.py', 'nonexistent.xlsx']):
        main()
        captured = capsys.readouterr()
        assert "not found" in captured.out.lower()


def test_main_function_extraction_error(capsys):
    """Test main function when extraction fails"""
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp.write(b"invalid excel")
        tmp.flush()
        tmp_path = tmp.name

    try:
        with patch('sys.argv', ['script.py', tmp_path]):
            main()
            captured = capsys.readouterr()
            # Check combined output (stdout + stderr) for maximum robustness
            # This works regardless of where error messages are printed
            combined_output = captured.out + captured.err
            assert "Traceback" in combined_output or len(captured.err) > 0
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def test_extract_bilancio_abbreviato_structure_completeness(temp_xlsx_abbreviato_file):
    """Test that all expected structure fields are present"""
    result = extract_bilancio_abbreviato_from_xlsx(temp_xlsx_abbreviato_file)
    
    # Check Stato Patrimoniale structure
    assert "Attivo" in result["Stato_patrimoniale"]
    assert "Passivo" in result["Stato_patrimoniale"]
    
    # Check Immobilizzazioni structure
    assert "Immobilizzazioni_Immateriali" in result["Stato_patrimoniale"]["Attivo"]["Immobilizzazioni"]
    assert "Immobilizzazioni_Materiali" in result["Stato_patrimoniale"]["Attivo"]["Immobilizzazioni"]
    assert "Immobilizzazioni_Finanziarie" in result["Stato_patrimoniale"]["Attivo"]["Immobilizzazioni"]
    
    # Check Conto Economico structure
    assert "Valore_della_produzione" in result["Conto_economico"]
    assert "Costi_di_produzione" in result["Conto_economico"]
    assert "Proventi_e_oneri_finanziari" in result["Conto_economico"]


def test_extract_bilancio_abbreviato_all_fields(temp_xlsx_abbreviato_file):
    """Test that all fields are correctly extracted"""
    result = extract_bilancio_abbreviato_from_xlsx(temp_xlsx_abbreviato_file)
    
    # Test Patrimonio Netto fields
    assert result["Stato_patrimoniale"]["Passivo"]["Patrimonio_netto"]["Capitale"] == 5000.0
    assert result["Stato_patrimoniale"]["Passivo"]["Patrimonio_netto"]["Riserve_da_sovrapprezzo_azioni"] == 1000.0
    assert result["Stato_patrimoniale"]["Passivo"]["Patrimonio_netto"]["Totale_patrimonio_netto"] == 9650.0
    
    # Test Conto Economico fields
    assert result["Conto_economico"]["Valore_della_produzione"]["Ricavi_delle_vendite_e_delle_prestazioni"] == 50000.0
    assert result["Conto_economico"]["Differenza_A_B"] == 10900.0
    assert result["Conto_economico"]["Risultato_prima_delle_imposte"] == 11740.0
    assert result["Conto_economico"]["Utile_(perdita)_dell_esercizio"] == 8390.0

