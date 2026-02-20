"""
Additional tests for balance_sheet_service.py focusing on Excel functionality
Extends test_balance_sheet_service.py with Excel-specific test cases
100% coverage target for Excel-related code paths
"""

import os
import sys
import tempfile
import pytest
from unittest.mock import patch, MagicMock, Mock
import openpyxl
from openpyxl import Workbook

# Add src to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

# Import the service module and test utilities
import importlib
service_module = importlib.import_module("src.app.api.v1.services.k_balance.balance_sheet_service")

from tests.test_balance_sheet_service import (
    SimpleUser, SimpleUserCompany, TbUserCompanyStub,
    KbaiBalanceStub, BalanceRecord, FakeFile, patch_service_dependencies
)


@pytest.fixture(autouse=True)
def patch_excel_dependencies():
    """Patch Excel-related dependencies"""
    original_extract_full = service_module.extract_bilancio_from_xlsx
    original_extract_abbr = service_module.extract_bilancio_abbreviato_from_xlsx
    original_detect_format = service_module.detect_excel_format
    original_extract_year = service_module.extract_balance_year
    
    # Default mocks
    service_module.extract_bilancio_from_xlsx = lambda x: {"test": "full"}
    service_module.extract_bilancio_abbreviato_from_xlsx = lambda x: {"test": "abbreviated"}
    service_module.detect_excel_format = lambda x: "full"
    service_module.extract_balance_year = lambda x: "2024-12-31"
    
    yield
    
    # Restore
    service_module.extract_bilancio_from_xlsx = original_extract_full
    service_module.extract_bilancio_abbreviato_from_xlsx = original_extract_abbr
    service_module.detect_excel_format = original_detect_format
    service_module.extract_balance_year = original_extract_year


@pytest.fixture
def temp_excel_file():
    """Create a temporary Excel file"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Bilancio 2024"
    ws['K4'] = 1000.0
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        yield tmp.name
    if os.path.exists(tmp.name):
        os.remove(tmp.name)


@pytest.fixture
def service_instance():
    return service_module.BalanceSheetService()


def test_balance_sheet_xlsx_full_format(service_instance, temp_excel_file):
    """Test balance sheet upload with XLSX full format"""
    service_module.detect_excel_format = lambda x: "full"
    service_module.extract_bilancio_from_xlsx = lambda x: {"assets": 1000}
    service_module.extract_balance_year = lambda x: "2024-12-31"
    
    fake_file = FakeFile(filename="report.xlsx")
    
    # Mock file.save to actually create the Excel file
    def mock_save(path):
        wb = Workbook()
        wb.save(path)
    
    fake_file.save = mock_save
    
    response, status = service_instance.balance_sheet(
        file=fake_file,
        company_id=1,
        year=2024,
        month=12,
        type="annual",
        mode="xlsx",
    )
    
    assert status == 201
    assert response["success"] is True
    if fake_file.saved_path and os.path.exists(fake_file.saved_path):
        os.remove(fake_file.saved_path)


def test_balance_sheet_xlsx_abbreviated_format(service_instance, temp_excel_file):
    """Test balance sheet upload with XLSX abbreviated format"""
    service_module.detect_excel_format = lambda x: "abbreviated"
    service_module.extract_bilancio_abbreviato_from_xlsx = lambda x: {"assets": 500}
    service_module.extract_balance_year = lambda x: "2024"
    
    fake_file = FakeFile(filename="report.xlsx")
    
    def mock_save(path):
        wb = Workbook()
        wb.save(path)
    
    fake_file.save = mock_save
    
    response, status = service_instance.balance_sheet(
        file=fake_file,
        company_id=1,
        year=2024,
        month=None,
        type="annual",
        mode="xlsx",
    )
    
    assert status == 201
    if fake_file.saved_path and os.path.exists(fake_file.saved_path):
        os.remove(fake_file.saved_path)


def test_balance_sheet_xlsx_unknown_format(service_instance, temp_excel_file):
    """Test balance sheet upload with unknown Excel format"""
    service_module.detect_excel_format = lambda x: "unknown"
    
    fake_file = FakeFile(filename="report.xlsx")
    
    def mock_save(path):
        wb = Workbook()
        wb.save(path)
    
    fake_file.save = mock_save
    
    response, status = service_instance.balance_sheet(
        file=fake_file,
        company_id=1,
        year=2024,
        month=12,
        type="annual",
        mode="xlsx",
    )
    
    assert status == 400
    # Service exposes the condition via the 'error' field and a detailed message.
    assert response.get("error") == "Unknown Excel format"
    assert "Could not determine matching script for this XLSX file" in response.get("message", "")
    if fake_file.saved_path and os.path.exists(fake_file.saved_path):
        os.remove(fake_file.saved_path)


def test_balance_sheet_xlsx_extraction_error(service_instance, temp_excel_file):
    """Test balance sheet upload when Excel extraction fails"""
    service_module.detect_excel_format = lambda x: "full"
    
    def failing_extract(x):
        raise ValueError("Extraction failed")
    
    service_module.extract_bilancio_from_xlsx = failing_extract
    
    fake_file = FakeFile(filename="report.xlsx")
    
    def mock_save(path):
        wb = Workbook()
        wb.save(path)
    
    fake_file.save = mock_save
    
    response, status = service_instance.balance_sheet(
        file=fake_file,
        company_id=1,
        year=2024,
        month=12,
        type="annual",
        mode="xlsx",
    )
    
    assert status == 500
    assert response["error"] == "Extraction error"
    if fake_file.saved_path and os.path.exists(fake_file.saved_path):
        os.remove(fake_file.saved_path)


def test_balance_sheet_xlsx_year_extraction_full_date(service_instance, temp_excel_file):
    """Test year extraction with full date format"""
    service_module.detect_excel_format = lambda x: "full"
    service_module.extract_bilancio_from_xlsx = lambda x: {"test": "data"}
    service_module.extract_balance_year = lambda x: "2024-12-31"
    
    fake_file = FakeFile(filename="report.xlsx")
    
    def mock_save(path):
        wb = Workbook()
        wb.save(path)
    
    fake_file.save = mock_save
    
    response, status = service_instance.balance_sheet(
        file=fake_file,
        company_id=1,
        year=2024,
        month=12,
        type="annual",
        mode="xlsx",
    )
    
    assert status == 201
    if fake_file.saved_path and os.path.exists(fake_file.saved_path):
        os.remove(fake_file.saved_path)


def test_balance_sheet_xlsx_year_extraction_year_only(service_instance, temp_excel_file):
    """Test year extraction with year only format"""
    service_module.detect_excel_format = lambda x: "full"
    service_module.extract_bilancio_from_xlsx = lambda x: {"test": "data"}
    service_module.extract_balance_year = lambda x: "2024"
    
    fake_file = FakeFile(filename="report.xlsx")
    
    def mock_save(path):
        wb = Workbook()
        wb.save(path)
    
    fake_file.save = mock_save
    
    response, status = service_instance.balance_sheet(
        file=fake_file,
        company_id=1,
        year=2024,
        month=None,
        type="annual",
        mode="xlsx",
    )
    
    assert status == 201
    if fake_file.saved_path and os.path.exists(fake_file.saved_path):
        os.remove(fake_file.saved_path)


def test_balance_sheet_xlsx_year_mismatch(service_instance, temp_excel_file):
    """Test year validation when extracted year doesn't match payload"""
    service_module.detect_excel_format = lambda x: "full"
    service_module.extract_bilancio_from_xlsx = lambda x: {"test": "data"}
    service_module.extract_balance_year = lambda x: "2023-12-31"
    
    fake_file = FakeFile(filename="report.xlsx")
    
    def mock_save(path):
        wb = Workbook()
        wb.save(path)
    
    fake_file.save = mock_save
    
    response, status = service_instance.balance_sheet(
        file=fake_file,
        company_id=1,
        year=2024,
        month=12,
        type="annual",
        mode="xlsx",
    )
    
    assert status == 400
    assert "Year mismatch" in response.get("message", "")
    if fake_file.saved_path and os.path.exists(fake_file.saved_path):
        os.remove(fake_file.saved_path)


def test_balance_sheet_xlsx_month_mismatch(service_instance, temp_excel_file):
    """Test month validation when extracted month doesn't match payload"""
    service_module.detect_excel_format = lambda x: "full"
    service_module.extract_bilancio_from_xlsx = lambda x: {"test": "data"}
    service_module.extract_balance_year = lambda x: "2024-11-30"
    
    fake_file = FakeFile(filename="report.xlsx")
    
    def mock_save(path):
        wb = Workbook()
        wb.save(path)
    
    fake_file.save = mock_save
    
    response, status = service_instance.balance_sheet(
        file=fake_file,
        company_id=1,
        year=2024,
        month=12,
        type="annual",
        mode="xlsx",
    )
    
    assert status == 400
    assert "Month mismatch" in response.get("message", "")
    if fake_file.saved_path and os.path.exists(fake_file.saved_path):
        os.remove(fake_file.saved_path)


def test_balance_sheet_xlsx_mode_validation(service_instance, temp_excel_file):
    """Test that xlsx mode only accepts xlsx files"""
    fake_file = FakeFile(filename="report.pdf")
    
    response, status = service_instance.balance_sheet(
        file=fake_file,
        company_id=1,
        year=2024,
        month=12,
        type="annual",
        mode="xlsx",
    )
    
    assert status == 400
    assert "Mode is set to" in response.get("message", "")


def test_balance_sheet_xls_mode_accepts_xlsx(service_instance, temp_excel_file):
    """Test that xls mode accepts xlsx files"""
    service_module.detect_excel_format = lambda x: "full"
    service_module.extract_bilancio_from_xlsx = lambda x: {"test": "data"}
    service_module.extract_balance_year = lambda x: "2024-12-31"
    
    fake_file = FakeFile(filename="report.xlsx")
    
    def mock_save(path):
        wb = Workbook()
        wb.save(path)
    
    fake_file.save = mock_save
    
    response, status = service_instance.balance_sheet(
        file=fake_file,
        company_id=1,
        year=2024,
        month=12,
        type="annual",
        mode="xls",
    )
    
    assert status == 201
    if fake_file.saved_path and os.path.exists(fake_file.saved_path):
        os.remove(fake_file.saved_path)


def test_balance_sheet_manual_mode_accepts_xlsx(service_instance, temp_excel_file):
    """Test that manual mode accepts xlsx files"""
    service_module.detect_excel_format = lambda x: "full"
    service_module.extract_bilancio_from_xlsx = lambda x: {"test": "data"}
    service_module.extract_balance_year = lambda x: "2024-12-31"
    
    fake_file = FakeFile(filename="report.xlsx")
    
    def mock_save(path):
        wb = Workbook()
        wb.save(path)
    
    fake_file.save = mock_save
    
    response, status = service_instance.balance_sheet(
        file=fake_file,
        company_id=1,
        year=2024,
        month=12,
        type="annual",
        mode="manual",
    )
    
    assert status == 201
    if fake_file.saved_path and os.path.exists(fake_file.saved_path):
        os.remove(fake_file.saved_path)


def test_extract_period_from_file_xlsx(service_instance, temp_excel_file):
    """Test _extract_period_from_file for XLSX files"""
    service_module.extract_balance_year = lambda x: "2024-12-31"
    
    year, month = service_instance._extract_period_from_file(temp_excel_file)
    assert year == 2024
    assert month == 12


def test_extract_period_from_file_xlsx_year_only(service_instance, temp_excel_file):
    """Test _extract_period_from_file for XLSX with year only"""
    service_module.extract_balance_year = lambda x: "2024"
    
    year, month = service_instance._extract_period_from_file(temp_excel_file)
    assert year == 2024
    assert month is None


def test_extract_period_from_file_xlsx_fallback(service_instance, temp_excel_file):
    """Test _extract_period_from_file fallback when extract_balance_year returns None"""
    service_module.extract_balance_year = lambda x: None
    
    # Should fallback to text extraction
    year, month = service_instance._extract_period_from_file(temp_excel_file)
    # May return None, None or extracted values
    assert year is None or isinstance(year, int)


def test_balance_sheet_xlsx_invalid_mode():
    """Test balance sheet upload with invalid mode"""
    service_instance = service_module.BalanceSheetService()
    fake_file = FakeFile(filename="report.xlsx")
    
    response, status = service_instance.balance_sheet(
        file=fake_file,
        company_id=1,
        year=2024,
        month=12,
        type="annual",
        mode="invalid_mode",
    )
    
    assert status == 400
    assert "Invalid mode" in response.get("message", "")


def test_balance_sheet_xlsx_file_extension_validation():
    """Test that only valid Excel extensions are accepted"""
    service_instance = service_module.BalanceSheetService()
    fake_file = FakeFile(filename="report.txt")
    
    response, status = service_instance.balance_sheet(
        file=fake_file,
        company_id=1,
        year=2024,
        month=12,
        type="annual",
        mode="manual",
    )
    
    assert status == 400
    assert "Only PDF, XLSX" in response.get("message", "")

