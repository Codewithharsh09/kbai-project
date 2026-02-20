import openpyxl
import re
from datetime import datetime

def detect_excel_format(file_path):
    """
    Detect whether the XLSX file belongs to Script 1 (Full format)
    or Script 2 (Abbreviated format) based on row signatures.
    """

    wb = openpyxl.load_workbook(file_path, data_only=True)
    try:
        ws = wb.active

        # Read text of row 4 (first significant label)
        row4_text = ""
        for cell in ws[4]:  # row index starts from 1, so this is row 4
            if cell.value:
                row4_text += str(cell.value).lower() + " "

        row4_text = row4_text.strip()

        # Matching Rules
        if "crediti verso soci" in row4_text or "parte richiamata" in row4_text:
            return "full"  # script.py format

        if "immobilizzazioni" in row4_text and "crediti verso soci" not in row4_text:
            return "abbreviated"  # script2.py format

        # Fallback: guess based on row count
        if ws.max_row > 120:
            return "full"
        else:
            return "abbreviated"
    finally:
        wb.close()

def extract_balance_year(file_path):
    """
    Extracts year or date from the header section (first 5 rows) of an XLSX balance file.
    Handles multiple formats like: 2023, 31/12/2023, al 31 dicembre 2023, Bilancio 2024, etc.
    """

    wb = openpyxl.load_workbook(file_path, data_only=True)
    try:
        ws = wb.active

        text_dump = ""

        # Read only first 5 rows (header region)
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if cell:
                    text_dump += str(cell).lower() + " "

        text_dump = text_dump.strip()

        # Regex patterns
        date_patterns = [
            r"\b(\d{1,2}[\/\.-]\d{1,2}[\/\.-]\d{4})\b",   # 31/12/2023 or 31-12-2023 or 31.12.2023
            r"\b(\d{4})\b"                               # 2023 (just year)
        ]

        # Italian month mapping for full date conversions
        months_it = {
            "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4, "maggio": 5, "giugno": 6,
            "luglio": 7, "agosto": 8, "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12
        }

        # 1) Numeric dates: dd/mm/yyyy, dd-mm-yyyy, dd.mm.yyyy
        m_num = re.search(date_patterns[0], text_dump)
        if m_num:
            value = m_num.group(1)
            try:
                normalized = value.replace("-", "/").replace(".", "/")
                dt = datetime.strptime(normalized, "%d/%m/%Y")
                return dt.strftime("%Y-%m-%d")  # standardized output
            except Exception:
                # Fall through to other strategies if parsing fails
                pass

        # 2) Italian text month format e.g. "al 31 dicembre 2023"
        for month in months_it:
            if month in text_dump and re.search(r"\b(19|20)\d{2}\b", text_dump):
                year = re.search(r"\b(19|20)\d{2}\b", text_dump).group(0)
                day_match = re.search(r"\b\d{1,2}\b", text_dump)
                day = int(day_match.group(0)) if day_match else 1
                return f"{year}-{months_it[month]:02d}-{day:02d}"

        # 3) Plain year-only fallback
        m_year = re.search(date_patterns[1], text_dump)
        if m_year:
            return m_year.group(1)

        return None  # no match found
    finally:
        wb.close()
